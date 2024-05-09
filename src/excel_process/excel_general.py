import openpyxl
from openpyxl.chart import LineChart, Reference
import pandas as pd
import matplotlib.pyplot as plt

class ExcelGeneralClass:
    def __init__(self, file, sheet = None):
        self.file = file
        self.sheet = sheet

    # 获取清单数据
    def read_data(self):
        if self.sheet is not None:
            data = pd.read_excel(self.file, sheet_name=self.sheet, keep_default_na=False)
            data_dict =data.to_dict('records')
            return data_dict
        else:
            raise ValueError("sheet 不能为空！")
        
    # 写入数据
    def write_data(self, data_dict):
        # 创建一个 ExcelWriter 对象
        writer = pd.ExcelWriter(self.file, engine='xlsxwriter')
        for item in data_dict:
            for sheet_name, sheet_data in item.items():
                if sheet_data:  # 检查字典中的数据是否为空
                    df = pd.DataFrame(sheet_data)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        # 保存 Excel 文件
        writer.close()
        
    # 写入图表
    def chart_data(self):
        if self.sheet is not None:
            wb = openpyxl.load_workbook(self.file)
            ws = wb[self.sheet]
            # 创建折线图
            chart = LineChart()
            chart.title = "计数统计"
            chart.x_axis.title = "Count_Name"
            chart.y_axis.title = "Count_Num"
            values = Reference(ws, min_col=3, min_row=1,max_col=3, max_row=ws.max_row)
            categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(values, titles_from_data=True)
            chart.set_categories(categories)
            ws.add_chart(chart, "F1")
            # 保存Excel文件
            wb.save(self.file)
        else:
            raise ValueError("sheet 不能为空！")
        
    def show_graph(self, count_name):
        graph_name = count_name + '-计数统计'
        df = pd.read_excel(self.file, sheet_name=self.sheet)
        plt.rcParams['font.sans-serif'] = ['SimHei']  # 设置中文字体为黑体
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号'-'显示为方块的问题
        # 绘制折线图
        plt.plot(df['count_name'], df['count_num'])
        plt.xlabel('Count_Name')
        plt.ylabel('Count_Num')
        plt.title(graph_name)
        plt.show()





